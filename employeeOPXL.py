import openpyxl as op
import os,sys,msvcrt as msv
FILENAME = "myemployee.xlsx"
if os.path.exists(FILENAME) == False:
    wb = op.Workbook()
    ws = wb.active
    ws['A1'] = "ID"
    ws['B1'] = "NAME"
    ws['C1'] = "DEPARTMENT"
    ws['D1'] = "D.O.B."
    ws['E1'] = "CITY"
    ws['F1'] = "PHONE NO."
    ws['G1'] = "SALARY"
    ws.title= "MainFile"
    fpsal = wb.create_sheet("EmpSalary")
    fpsal['A1'] = "ID"
    fpsal['B1'] = "BASIC"
    fpsal['C1'] = "H.R.A"
    fpsal['D1'] = "D.A"
    fpsal['E1'] = "INCOME TAX"
    fpsal['F1'] = "NET SALARY"
    wb.save(FILENAME)
    wb.close()

def insertdata():
    os.system("cls")
    print("\nInserting the data module\n\n")
    fp = op.load_workbook(FILENAME)
    ws = fp.active
    ids = (cell.value for cell in ws['A'][1:] if cell.value is not None)
    if ws.max_row == 1:
        id = 1001
    else:
        id = max(ids) + 1 
    print(f"ID of employee -- {id}")
    name = str(input("Enter Name of employee -- "))
    dept = str(input("Enter Department of employee -- "))
    doj = str(input("Enter the Date of Joining as (DD/MM/YYYY) of employee -- "))
    city = str(input("Enter City of employee -- "))
    mob = str(input("Enter Phone number of employee -- "))
    salary = str(input("Enter Salary of employee -- "))
    data = [id,name,dept,doj,city,mob,salary]
    ws.append(data)
    row = 2
    fpws = fp["EmpSalary"]
    sal = int(salary)
    HRA = int((int(salary)*(0.2)))
    DA = int((int(salary)*(0.6)))
    INTX = int((int(salary) + HRA + DA)*0.15)
    data = [id,salary, HRA, DA,INTX, (sal + HRA + DA)-INTX]
    fpws.append(data)
    fp.save(FILENAME)
    fp.close()

def showdata():
    os.system("cls")
    print("\nShowing the data module\n\n")
    fp = op.load_workbook(FILENAME)
    ws = fp.active
    print("-"*72)
    for i in ws.iter_rows(values_only=True):  
        print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s"% i)
        print("-"*72)
    fp.save(FILENAME)
    fp.close()

def searchdata():
    os.system("cls")
    print("\nSearching data module\n\n")
    
    fp = op.load_workbook(FILENAME)
    ws = fp.active
    
    print("1 -> Search by ID      2 -> Search by Department      3 -> Search by City")
    c = msv.getch().decode()
    
    if c == '1':
        search = int(input("\nEnter the ID to search employee -- "))
        found = 0
        
        for i in ws.iter_rows(values_only=True):
            if i[0] == "ID":
                continue
            
            if i[0] == search:
                print("-"*72)
                print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % i)
                print("-"*72)
                found = 1
                break
            
        if found == 0:
            print(f"\nOOPS!!! No record found with ID '{search}'")
            
    elif c == '2' or c == '3':
        
        if c == '2':
            search = input("\nEnter the Department -- ")
            index = 2
            label = "Department"
        else:
            search = input("\nEnter the City -- ")
            index = 4
            label = "City"
            
        print("-"*72)
        matched = []
        
        for i in ws.iter_rows(values_only=True):
            if i[0] == "ID":
                continue
            
            if i[index].lower() == search.lower():
                print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % i)
                print("-"*72)
                matched.append(i)
                
        if len(matched) == 0:
            print(f"\n{label} not found\n")
            
        else:
            id = int(input(f"\nEnter ID of Employee from {label} '{search}' -- "))
            found = 0
            
            for j in matched:
                if j[0] == id:
                    print("-"*72)
                    print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % j)
                    print("-"*72)
                    found = 1
                    break
                
            if found == 0:
                print(f"\nID - {id} not found in {label} {search}\n")
    
    fp.close()

def updatedata():
    os.system("cls")
    print("\nUpdating the data module\n\n")
    
    fp = op.load_workbook(FILENAME)
    ws = fp.active
    
    id = int(input("\nEnter ID of the employee you wanna update -- "))
    found = 0
    rows = 1
    
    print("-"*72)
    for i in ws.iter_rows(values_only=True):
        if i[0] == id:
            print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % i)
            print("-"*72)
            found = 1
            break
        rows += 1
    if found == 0:
        print("\nOOPS!! RECORD NOT FOUND!! \n")
    else:
        print("\nWhat do you want to update?")
        print("""1 -> Name
2 -> Department  
3 -> Date of Joining  
4 -> City  
5 -> Phone  
6 -> Salary  
7 -> All
""")
        c = msv.getch().decode()
        update = 0
        if c == '1':
            ws.cell(rows,2).value = input("\nEnter new name -- ")
            update = 1
        elif c == '2':
            ws.cell(rows,3).value = input("\nEnter new Department -- ")
            update = 1
        elif c == '3':
            ws.cell(rows,4).value = input("\nEnter new Date of Joining -- ")
            update = 1
        elif c == '4':
            ws.cell(rows,5).value = input("\nEnter new City -- ")
            update = 1
        elif c == '5':
            ws.cell(rows,6).value = input("\nEnter new Phone No. -- ")
            update = 1
        elif c == '6':
            ws.cell(rows,7).value = input("\nEnter new Salary (in ₹) -- ")
            update = 1
        elif c == '7':
            ws.cell(rows,2).value = input("\nEnter new name -- ")
            ws.cell(rows,3).value = input("\nEnter new Department -- ")
            ws.cell(rows,4).value = input("\nEnter new Date of Joining -- ")
            ws.cell(rows,5).value = input("\nEnter new City -- ")
            ws.cell(rows,6).value = input("\nEnter new Phone No. -- ")
            ws.cell(rows,7).value = input("\nEnter new Salary (in ₹) -- ")
            update = 1
        else:
            print("\nWrong option chosen !!\n")
    if update == 0:
        print("\nError in updating, Kindly Retry!!!\n")
    elif update == 1:
        print("\nData successfully Updated")
    fp.save(FILENAME)
    fp.close()

def deletedata():
    os.system("cls")
    print("\nDeleting data module\n\n")
    
    fp = op.load_workbook(FILENAME)
    ws = fp.active
    delWS = fp["deletedEmp"] if "deletedEmp" in fp.sheetnames else fp.create_sheet("deletedEmp")
    
    print("1 -> Temporary Delete      2 -> Permanent Delete")
    ch = msv.getch().decode()
    
    search = int(input("\nEnter ID to delete -- "))
    found = 0
    rows = 1
    
    for i in ws.iter_rows(values_only=True):
        if i[0] == "ID":
            rows += 1
            continue
        
        if i[0] == search:
            print("-"*72)
            print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % i)
            print("-"*72)
            found = 1
            break
        rows += 1
    
    if found == 0:
        print("\nNo record found")
        fp.close()
        return
    print("\nAre you sure, press 1 else any key\n")
    confirm = msv.getch().decode()
    
    if confirm.lower() != '1':
        print("\nDeletion cancelled")
        fp.close()
        return
    
    if ch == '1':
        delWS.append(i)
    ws.delete_rows(rows)
    
    fp.save(FILENAME)
    fp.close()
    
    print("\nRecord deleted successfully")

def recoverdata():
    os.system("cls")
    print("\nRecovering the deleted records module...\n\n")
    
    fp = op.load_workbook(FILENAME)
    fpws = fp.active
    
    fpdel = fp["deletedEmp"] if "deletedEmp" in fp.sheetnames else fp.create_sheet("deletedEmp")
    
    print("Temporarily Deleted Records are given below -- \n")
    print("-"*72)
    for i in fpdel.iter_rows(values_only=True):  
        print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s"% i)
        print("-"*72)
    
    print("\n1 -> Recover 1 ID     2 -> Recover All")
    c = msv.getch().decode()
    if c == '1':
        row = 1
        id = int(input("\nEnter ID to recover -- "))
        print("\n" + ("-"*72))
        for j in fpdel.iter_rows(values_only=True):
            if j[0] == id:
                print("%-5s %-10s %-11s %-13s %-10s %-11s %-7s" % j)
                print("-"*72)
                break
            else:
                print("\nData not found\n")
            row += 1
        print("\nAre you sure you wanna recover, if yes press 1 else any key\n")
        ch = msv.getch().decode()
        if ch != '1':
            print("\nData not Recovered\n")
            fp.save(FILENAME)
            fp.close()
            return
        elif ch == '1':
            fpws.append(j)
        fpdel.delete_rows(row)
    
    elif c == '2':
        print("\nAre you sure you wanna recover all, if yes press 1 else any key\n")
        ch = msv.getch().decode()
        
        if ch != '1':
            print("\nData not Recovered\n")
        elif ch == '1':
            for j in fpdel.iter_rows(values_only=True):
                fpws.append(j)
            fpdel.delete_rows(1, fpdel.max_row)
            print("\nAll records recovered successfully\n")
    
    fp.save(FILENAME)
    fp.close()


def salarytell():
    os.system("cls")
    print("\nSalary of Employees module....\n\n")
    fp = op.load_workbook(FILENAME)
    ws = fp["EmpSalary"]
    
    id = int(input("Enter the ID of employee whose salary report you want -- "))
    for i in ws.iter_rows(values_only=True):
        if i[0] == id:
            print("\nAre you sure, press 1 else any key")
            c = msv.getch().decode()
            if c == '1':
                print(f"""ID         --- {id}
Basic      --- ₹{i[1]}
H.R.A      --- ₹{i[2]}
D.A        --- ₹{i[3]}
Income Tax --- ₹{i[4]}

Net Salary --- ₹{i[5]}
""")
    fp.close()

def finalcard():
    os.system("cls")
    
    fp = op.load_workbook(FILENAME)
    fpws = fp["MainFile"]
    fpsal = fp["EmpSalary"]
    
    id = int(input("\nEnter ID of the employee -- "))
    
    for i in fpws.iter_rows(values_only=True):
        if i[0] == "ID":
            continue
        if i[0] == id:
            name = i[1]
            dept = i[2]
            doj = i[3]
            city = i[4]
            PhNo = i[5]
            sal = i[6]
    
    for j in fpsal.iter_rows(values_only=True):
        if j[0] == id:
            hra = j[2]
            da = j[3]
            inctax = j[4]
            netsal = j[5]
    
    print("\n\n")
    print("="*35)
    print("NORTH WESTERN RAILWAY".center(35))
    print("="*35)
    
    print(f"""ID              --  {id}
Name            --  {name}
Department      --  {dept}
Date of Joining --  {doj}
City            --  {city}
Phone No.       --  {PhNo}""")
    print("-"*35)
    print(f"""Basic           --  ₹{sal}
H.R.A           --  ₹{hra}
D.A             --  ₹{da}
Income Tax      --  ₹{inctax}

Net Salary      --  ₹{netsal}""")
    
    print("="*35)
    print("\n\n")
    
    print("Press 1 to create PDF else any key")
    c = msv.getch().decode()
    
    if c == '1':
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        
        doc = SimpleDocTemplate(str(id)+str(name)+".pdf")
        styles = getSampleStyleSheet()
        content = []
        
        # Image
        img = Image(r"rlogo.jpg", width=120, height=110)
        content.append(img)
        content.append(Spacer(1, 20))
        
        # Heading
        heading = Paragraph(
            "<b><font size=16>NORTH WESTERN RAILWAY</font></b>",
            styles["Title"]
        )
        content.append(heading)
        
        # ➖ Line
        line = Table([[""]], colWidths=[400])
        line.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,-1), 1, colors.black),
        ]))
        content.append(line)
        content.append(Spacer(1, 10))
        
        # Employee Table
        data1 = [
            ["ID", "--", id],
            ["Name", "--", name],
            ["Department", "--", dept],
            ["Date of Joining", "--", str(doj)],
            ["City", "--", city],
            ["Phone No.", "--", PhNo],
        ]
        
        table1 = Table(data1, colWidths=[150, 50, 200])
        table1.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 11),
        ]))
        content.append(table1)
        content.append(Spacer(1, 10))
        
        # Line
        line2 = Table([[""]], colWidths=[400])
        line2.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,-1), 1, colors.black),
        ]))
        content.append(line2)
        
        content.append(Spacer(1, 10))
        
        # Salary Table
        data2 = [
            ["Basic", "--", f"Rs. {sal}"],
            ["H.R.A", "--", f"Rs. {hra}"],
            ["D.A", "--", f"Rs. {da}"],
            ["Income Tax", "--", f"Rs. {inctax}"],
            ["", "", ""],
            ["Net Salary", "--", f"Rs. {netsal}"],
        ]
        
        table2 = Table(data2, colWidths=[150, 50, 200])
        table2.setStyle(TableStyle([
            ('FONTSIZE', (0,0), (-1,-1), 11),
            ('LINEABOVE', (0,5), (-1,5), 1, colors.black),
        ]))
        content.append(table2)
        
        content.append(Spacer(1, 10))
        
        # Bottom Line
        line3 = Table([[""]], colWidths=[400])
        line3.setStyle(TableStyle([
            ('LINEABOVE', (0,0), (-1,-1), 1, colors.black),
        ]))
        content.append(line3)
        
        # Build PDF
        doc.build(content)
        
        print("\nPDF created successfully!\n")
        
    fp.close()

def MainModule():
    os.system("cls")                                                
    print("Employee File number 2....  \n")
    print("""1 -> Insert records in the file
2 -> Show the records in the file
3 -> Search data in the file
4 -> Update the records of the file
5 -> Delete any record in the file
6 -> Recover deleted records
7 -> Salary calculator of employees
8 -> Get full record of any employee
Press any key other than 1-8 for "EXIT" 
""")
    c = msv.getch().decode()
    if c == '1':
        insertdata()
    elif c == '2':
        showdata()
    elif c == '3':
        searchdata()
    elif c == '4':
        updatedata()
    elif c == '5':
        deletedata()
    elif c == '6':
        recoverdata()
    elif c == '7':
        salarytell()
    elif c == '8':
        finalcard()
    else:
        print("\nTHANKS FOR USING OUR MODULE.... \n")
        sys.exit(0)
    print("\nDo you want to continue... press 1 else any key.... ")
    ch = msv.getch().decode()
    if ch == '1':
        MainModule()
    else:
        print("\nTHANKS FOR USING OUR MODULE.... \n")
        sys.exit(0)
MainModule()